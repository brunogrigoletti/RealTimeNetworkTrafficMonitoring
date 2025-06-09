import datetime
import struct
import sys
import threading
import os
import time
import signal
from winpcapy import WinPcapUtils, WinPcapDevices
from openpyxl import Workbook, load_workbook

excel_lock = threading.Lock()

is_shutting_down = False

LAYER2_LOG = "layer2.xlsx"
LAYER3_LOG = "layer3.xlsx"
LAYER4_LOG = "layer4.xlsx"

def delete_existing_logs():
    log_files = [LAYER2_LOG, LAYER3_LOG, LAYER4_LOG]
    print("Checking for existing log files...")
    for file in log_files:
        try:
            if os.path.exists(file):
                os.remove(file)
                print(f"Deleted existing log file: {file}")
        except Exception as e:
            print(f"Error deleting {file}: {e}")

def write_excel_log(filename, header, row):
    with excel_lock:
        file_exists = os.path.exists(filename) and os.path.getsize(filename) > 0
        try:
            if file_exists:
                wb = load_workbook(filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(header)
            ws.append(row)
            wb.save(filename)
        except Exception as e:
            print(f"Error on file '{filename}': {e}")

def mac_addr(bytes_addr):
    # Converte endereço MAC em bytes para String legível
    return ':'.join('%02x' % b for b in bytes_addr)

# Inicialização dos contadores de pacotes
eth_count = 0
ip4_count = 0
ip6_count = 0
arp_count = 0
tcp_count = 0
udp_count = 0

def packet_callback(win_pcap, param, header, pkt_data):
    global eth_count, ip4_count, ip6_count, arp_count, tcp_count, udp_count

    # Obtém o timestamp do header
    ts_sec = header.contents.ts.tv_sec
    ts_usec = header.contents.ts.tv_usec
    timestamp = datetime.datetime.fromtimestamp(ts_sec + ts_usec / 1_000_000)

    # Extrai e processa o cabeçalho Ethernet do pacote capturado
    eth_header = pkt_data[:14]
    if len(eth_header) < 14:
        return
    
    # Desempacota o cabeçalho Ethernet
    # '!6s': 6 bytes para o endereço MAC de destino
    # '6s': 6 bytes para o endereço MAC de origem
    # 'H': 2 bytes para o campo "EtherType" (protocolo)
    # '!': ordem de bytes network (big-endian)
    eth = struct.unpack('!6s6sH', eth_header)
    eth_count += 1

    # Dicionário EtherType para nome do protocolo
    ethertype_names = {
        0x0800: "IPv4",
        0x0806: "ARP",
        0x86DD: "IPv6",
    }
    proto_hex = eth[2]
    proto_name = ethertype_names.get(proto_hex, "Unknown")

    # Imprime o cabeçalho Ethernet e escreve no log do Excel
    # print_ethernet_header(timestamp, eth, proto_name, header.contents.len, eth_count)
    write_excel_log(
        LAYER2_LOG,
        ['Timestamp', 'Source MAC', 'Destination MAC', 'Protocol', 'Protocol Name', 'Length'],
        [timestamp.strftime('%d/%m/%Y %I:%M:%S %p'), mac_addr(eth[1]), mac_addr(eth[0]), f"{eth[2]:#06x}", proto_name, header.contents.len]
    )

    # Extrai e processa o cabeçalho IPv4 do pacote capturado
    ip4_header = pkt_data[14:34]
    if len(ip4_header) < 20:
        return
    
    # Desempacota o cabeçalho IPv4
    # '!': ordem de bytes network (big-endian)
    # 'B': 1 byte para Versão + IHL (Internet Header Length)
    # 'B': 1 byte para Type of Service (ToS)
    # 'H': 2 bytes para Total Length do pacote
    # 'H': 2 bytes para Identification
    # 'H': 2 bytes para Flags + Fragment Offset
    # 'B': 1 byte para Time To Live (TTL)
    # 'B': 1 byte para Protocol (identificador)
    # 'H': 2 bytes para Header Checksum
    # '4s': 4 bytes para o IP de origem
    # '4s': 4 bytes para o IP de destino
    ipv4 = struct.unpack('!BBHHHBBH4s4s', ip4_header)

    src_ip4 = '.'.join(map(str, ipv4[8]))
    dst_ip4 = '.'.join(map(str, ipv4[9]))
 
    # Extrai e processa o cabeçalho IPv6 do pacote capturado
    ip6_header = pkt_data[14:54]
    if len(ip6_header) < 40:
        return
    
    # Desempacota o cabeçalho IPv6
    # '!': ordem de bytes network (big-endian)
    # 'I': 4 bytes para Versão, Traffic Class e Flow Label
    # 'H': 2 bytes para Payload Length
    # 'B': 1 byte para Next Header (identificador)
    # 'B': 1 byte para Hop Limit (TTL do IPv6)
    # '16s': 16 bytes para o IP de origem
    # '16s': 16 bytes para o IP de destino
    ipv6 = struct.unpack('!IHBB16s16s', ip6_header)

    src_ip6 = ':'.join(f"{ipv6[4][i]<<8 | ipv6[4][i+1]:x}" for i in range(0,16,2))
    dst_ip6 = ':'.join(f"{ipv6[5][i]<<8 | ipv6[5][i+1]:x}" for i in range(0,16,2))

    # Extrai e processa o cabeçalho ARP do pacote capturado
    arp_header = pkt_data[14:42]
    if len(arp_header) < 28:
        return
    
    # Desempacota o cabeçalho ARP
    # '!': ordem de bytes network (big-endian)
    # 'H': 2 bytes para Hardware Type
    # 'H': 2 bytes para Protocol Type
    # 'B': 1 byte para Hardware Size
    # 'B': 1 byte para Protocol Size
    # 'H': 2 bytes para Opcode (Operação ARP)
    # '6s': 6 bytes para o MAC de Origem
    # '4s': 4 bytes para o IP de Origem (IPv4)
    # '6s': 6 bytes para o MAC de Destino
    # '4s': 4 bytes para o IP de Destino (IPv4)
    arp = struct.unpack('!HHBBH6s4s6s4s', arp_header)

    src_arp = '.'.join(map(str, arp[6]))
    dst_arp = '.'.join(map(str, arp[8]))

    # Imprime o cabeçalho IP e escreve no log do Excel
    # print_ipv4_header(timestamp, src_ip4, dst_ip4, ipv4, header.contents.len, ip4_count)
    # print_ipv6_header(timestamp, src_ip6, dst_ip6, ipv6, header.contents.len, ip6_count)
    # print_arp_header(timestamp, src_arp, dst_arp, arp, header.contents.len, arp_count)
    if proto_name == "IPv4":
        ip4_count += 1

        write_excel_log(
            LAYER3_LOG,
            ['Timestamp', 'Protocol', 'Source IP', 'Destination IP', 'Protocol ID', 'Length'],
            [timestamp.strftime('%d/%m/%Y %I:%M:%S %p'), proto_name, src_ip4, dst_ip4, ipv4[6], header.contents.len]
        )
    elif proto_name == "IPv6":
        ip6_count += 1

        write_excel_log(
            LAYER3_LOG,
            ['Timestamp', 'Protocol', 'Source IP', 'Destination IP', 'Protocol ID', 'Length'],
            [timestamp.strftime('%d/%m/%Y %I:%M:%S %p'), proto_name, src_ip6, dst_ip6, ipv6[2], header.contents.len]
        )
    elif proto_name == "ARP":
        arp_count += 1

        write_excel_log(
            LAYER3_LOG,
            ['Timestamp', 'Protocol', 'Source IP', 'Destination IP', 'Protocol ID', 'Length'],
            [timestamp.strftime('%d/%m/%Y %I:%M:%S %p'), proto_name, src_arp, dst_arp, arp[4], header.contents.len]
        )

    # Extrai e processa o cabeçalho de transporte do pacote capturado
    transport_header = None
    # TCP
    if ipv4[6] == 6:
        transport_header = pkt_data[34:54]

        # Desempacota o cabeçalho de transporte TCP
        # '!': ordem de bytes network (big-endian)
        # 'H': 2 bytes para Porta de Origem
        # 'H': 2 bytes para Porta de Destino
        # 'L': 4 bytes para Número de Sequência
        # 'L': 4 bytes para Número de Acknowledgment
        # 'B': 1 byte para Data Offset + Reserved + Flags
        # 'B': 1 byte para Janela (Window Size)
        # 'H': 2 bytes para Checksum
        # 'H': 2 bytes para Urgent Pointer
        # 'H': 2 bytes para Opções (se houver)
        tcp = struct.unpack('!HHLLBBHHH', transport_header)
        
    # UDP
    elif ipv6[2] == 17:
        transport_header = pkt_data[34:42]

        # Desempacota o cabeçalho de transporte UDP
        # '!': ordem de bytes network (big-endian)
        # 'H': 2 bytes para Porta de Origem
        # 'H': 2 bytes para Porta de Destino
        # 'H': 2 bytes para Comprimento
        # 'H': 2 bytes para Checksum
        udp = struct.unpack('!HHHH', transport_header) 

    if transport_header and ipv4[6] == 6:
        proto = 'TCP'
        src_ip_str = src_ip4
        src_port = tcp[0]
        dst_ip_str = dst_ip4
        dst_port = tcp[1]
        pkt_count = tcp_count
    elif transport_header and ipv6[2] == 17:
        proto = 'UDP'
        src_ip_str = src_ip6
        src_port = udp[0]
        dst_ip_str = dst_ip6
        dst_port = udp[1]
        pkt_count = udp_count
    else:
        proto = 'Unknown'
        src_ip_str = 'N/A'
        src_port = 'N/A'
        dst_ip_str = 'N/A'
        dst_port = 'N/A'
        pkt_count = 'N/A'
    
    # Imprime o cabeçalho UDP/TCP e escreve no log do Excel
    # print_transport_header(timestamp,proto,src_ip_str,src_port,dst_ip_str,dst_port,header.contents.len,pkt_count)
    if transport_header and proto == 'TCP':
        tcp_count += 1
        
        write_excel_log(
            LAYER4_LOG,
            ['Timestamp', 'Protocol', 'Source IP', 'Source Port', 'Destination IP', 'Destination Port', 'Length'],
            [timestamp.strftime('%d/%m/%Y %I:%M:%S %p'), proto, src_ip_str, src_port, dst_ip_str, dst_port, header.contents.len]
        )
    elif transport_header and proto == 'UDP':
        udp_count += 1

        write_excel_log(
            LAYER4_LOG,
            ['Timestamp', 'Protocol', 'Source IP', 'Source Port', 'Destination IP', 'Destination Port', 'Length'],
            [timestamp.strftime('%d/%m/%Y %I:%M:%S %p'), proto, src_ip_str, src_port, dst_ip_str, dst_port, header.contents.len]
        )
    
def print_ethernet_header(timestamp, eth, proto_name, header_len, eth_count):
    print("=" * 50)
    print("\nEthernet Header:")
    print(f"  Timestamp: {timestamp.strftime('%d/%m/%Y %I:%M:%S %p')}")
    print(f"  Source MAC: {mac_addr(eth[1])}")
    print(f"  Destination MAC: {mac_addr(eth[0])}")
    print(f"  Protocol: {eth[2]:#06x} - {proto_name}")
    print(f"  Length: {header_len} bytes")
    print(f"  Packet Count: {eth_count}")

def print_ipv4_header(timestamp, src_ip4, dst_ip4, ipv4, header_len, ip4_count):
    print("\nIPv4 Header:")
    print(f"  Timestamp: {timestamp.strftime('%d/%m/%Y %I:%M:%S %p')}")
    print(f"  Source IP: {src_ip4}")
    print(f"  Destination IP: {dst_ip4}")
    print(f"  Protocol ID: {ipv4[6]}")
    print(f"  Length: {header_len} bytes")
    print(f"  Packet Count: {ip4_count}")

def print_ipv6_header(timestamp, src_ip6, dst_ip6, ipv6, header_len, ip6_count):
    print("\nIPv6 Header:")
    print(f"  Timestamp: {timestamp.strftime('%d/%m/%Y %I:%M:%S %p')}")
    print(f"  Source IP: {src_ip6}")
    print(f"  Destination IP: {dst_ip6}")
    print(f"  Protocol ID: {ipv6[2]}")
    print(f"  Length: {header_len} bytes")
    print(f"  Packet Count: {ip6_count}")

def print_arp_header(timestamp, src_arp, dst_arp, arp, header_len, arp_count):
    print("\nARP Header:")
    print(f"  Timestamp: {timestamp.strftime('%d/%m/%Y %I:%M:%S %p')}")
    print(f"  Source IP: {src_arp}")
    print(f"  Destination IP: {dst_arp}")
    print(f"  Protocol ID: {arp[4]}")
    print(f"  Length: {header_len} bytes")
    print(f"  Packet Count: {arp_count}")

def print_transport_header(timestamp, proto, src_ip, src_port, dst_ip, dst_port, header_len, count):
    print("\nTransport Header:")
    print(f"  Timestamp: {timestamp.strftime('%d/%m/%Y %I:%M:%S %p')}")
    print(f"  Protocol: {proto}")
    print(f"  Source IP: {src_ip}")
    print(f"  Source Port: {src_port}")
    print(f"  Destination IP: {dst_ip}")
    print(f"  Destination Port: {dst_port}")
    print(f"  Length: {header_len} bytes")
    print(f"  Packet Count: {count}")
    print("=" * 50)

def list_interfaces():
    # Lista todas as interfaces disponíveis para captura
    print("\nAvailable interfaces:")
    with WinPcapDevices() as devices:
        device_list = []
        for i, device in enumerate(devices):
            desc = device.description.decode() if isinstance(device.description, bytes) else device.description
            desc = desc if desc else "No description available"
            print(f"{i}: {desc}")
            device_list.append(desc)

        # Escolha da interface pelo índice
        idx = int(input("\nType the interface number to start capturing: "))
        return device_list[idx]

stop_event = threading.Event()

def capture_packets(interface):
    # Captura dos pacotes na interface escolhida
    try:
        WinPcapUtils.capture_on(interface, packet_callback)
    except Exception as e:
        print(f"Capture error: {e}")

def loading_spinner(stop_event):
    spinner = ['|', '/', '-', '\\']
    idx = 0
    while not stop_event.is_set():
        sys.stdout.write(f"\rCapturing... {spinner[idx % len(spinner)]}")
        sys.stdout.flush()
        idx += 1
        time.sleep(0.1)
    
    sys.stdout.write('\r' + ' ' * 40 + '\r\n')
    sys.stdout.flush()

def graceful_shutdown(signum=None, frame=None):
    global is_shutting_down
    
    if is_shutting_down:
        return
    
    is_shutting_down = True
    print("\n\nCapture stopped...")
    
    stop_event.set()
    
    time.sleep(0.5)
    
    if eth_count > 0:
        print("=" * 50)
        print(f"Summary:")
        print(f"Ethernet frames: {eth_count}")
        print(f"IPv4 packets:    {ip4_count}")
        print(f"IPv6 packets:    {ip6_count}")
        print(f"ARP packets:     {arp_count}")
        print(f"TCP segments:    {tcp_count}")
        print(f"UDP datagrams:   {udp_count}")
        print("=" * 50)
    
    print("Shutdown complete. Excel files saved safely.")
    sys.exit(0)

def main():
    signal.signal(signal.SIGINT, graceful_shutdown)
    signal.signal(signal.SIGTERM, graceful_shutdown)

    try:
        delete_existing_logs()

        interface = list_interfaces()
        print(f"\nSelected interface: '{interface}'\n")
        capture_thread = None
        spinner_thread = None

        while True:
            cmd = input("Type 'cat' to start capturing or 'stop' to stop: ").strip().lower()
            if cmd == "cat":
                if capture_thread and capture_thread.is_alive():
                    print("Capture is already running!")
                else:
                    stop_event.clear()
                    capture_thread = threading.Thread(target=capture_packets, args=(interface,), daemon=True)
                    spinner_thread = threading.Thread(target=loading_spinner, args=(stop_event,), daemon=True)
                    capture_thread.start()
                    spinner_thread.start()  
            elif cmd == "stop":
                if capture_thread and capture_thread.is_alive():
                    stop_event.set()

                    if spinner_thread and spinner_thread.is_alive():
                        spinner_thread.join(timeout=1.0)
                    
                    graceful_shutdown()
                else:
                    print("No capture is running!")
            else:
                print("Unknown command!")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        graceful_shutdown()

if __name__ == "__main__":
    main()